import openpyxl
import csv

from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active


with open('task_17.csv') as f:
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        ws.append(row)

column_a = ws['A']
column_b = ws['B']
column_c = ws['C']
column_d = ws['D']

row_1 = ws['1']
row_2 = ws['2']
row_3 = ws['3']
row_4 = ws['4']

column_a, row_1 = row_1, column_a
column_b, row_2 = row_2, column_b
column_c, row_3 = row_3, column_c
column_d, row_4 = row_4, column_d

ws.append(row_1)
ws.append(row_2)
ws.append(row_3)
ws.append(row_4)

ws.delete_rows(idx=1, amount=7)  # тут видаляю всі пусті рядки, інакше якісь залишки
# фалів призводять до можливості читання лише відновленого файлу Excel

ws.delete_rows(idx=3, amount=1)
ws.insert_rows(idx=1, amount=1)

pers_nums = ['', 'Person 1', 'Person 2', 'Person 3', 'Person 4', 'Person 5', 'Person 6']
for i in range(1, 7):
    cellref = ws.cell(row=1, column=i+1)
    cellref.value = pers_nums[i]

dim_holder = DimensionHolder(worksheet=ws)

for col in range(ws.min_column, ws.max_column + 1):
    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

ws.column_dimensions = dim_holder

wb.save('file2.xlsx')
