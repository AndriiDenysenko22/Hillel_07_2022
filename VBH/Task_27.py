import openpyxl
import unittest

from openpyxl import utils
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import datetime
from datetime import datetime

from Cl_person import Person

wb2 = openpyxl.Workbook()
work_file = input('Please, enter name of new data file')
wb2.save(work_file+'.xlsx')

while True:
    print('4: Find in file')
    print('3: Save file')
    print('2: Input person')
    print('1: Add from file')
    print('0: Exit')

    ch = int(input('Select operation: '))

    if ch == 4:
        wb = openpyxl.load_workbook(work_file+'.xlsx')
        ws = wb.active

        def wordfinder(searchString):
            find_pers = []
            for i in range(1, ws.max_row + 1):
                for j in range(1, ws.max_column + 1):
                    try:
                        if searchString in ws.cell(i, j).value:
                            ws._current_row = i
                            for j in range(1, ws.max_column + 1):
                                find_pers.append(ws.cell(ws._current_row, j).value)
                                chunks = [find_pers[x:x + 6] for x in range(0, len(find_pers), 6)]
                    except TypeError:
                        pass
            for item in chunks:
                item4 = datetime.strptime(item[4], '%d.%m.%Y')
                if item[5] is None:
                    today = datetime.today()
                else:
                    item5 = datetime.strptime(item[5], '%d.%m.%Y')
                    today = item5
                try:
                    birthday = item4.replace(year=today.year)

                except ValueError:
                    birthday = item4.replace(year=today.year, month=item4.month + 1, day=1)

                if birthday > today:
                    aaa = today.year - item4.year - 1
                else:
                    aaa = today.year - item4.year

                item.append(f'age: {aaa} years')
                # if item[5] is None:
                #     print(f'{item[2]} {item[0]} {item[1]}, sex: {item[3]}, birthdate: {item[4]}, {item[6]}', sep='\n')
                # else:
                print(f'{item[1]} {item[0]} {item[2]}, sex: {item[3]}, birthdate: {item[4]}, '
                      f'date of death: {item[5]}, {item[6]}', sep='\n')


        find = input('Enter the word, you wish to find: ')
        wordfinder(find)


    if ch == 3:
        wb2.save(work_file + '.xlsx')


    if ch == 2:
        user_input = (Person.get_list('first name', 'sex', 'date of birth'))
        ws2 = wb2.active
        mx = ws2.max_row
        mxc = ws2.max_column
        if mx == 1:
            first_row = ['First Name', 'Last Name', 'Patronimic', 'Sex', 'Date of birth', 'Date of death']
            ws2.append(first_row)


        for column, entry in enumerate(user_input, start=1):
            ws2.cell(row=mx+1, column=column, value=entry)
            dim_holder = DimensionHolder(worksheet=ws2)
            for col in range(ws2.min_column, ws2.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(ws2, min=col, max=col, width=20)
            ws2.column_dimensions = dim_holder


    if ch == 1:
        name_of_file = input('Please, enter a file name')
        try:
            wb1 = openpyxl.load_workbook(name_of_file)
        except openpyxl.utils.exceptions.InvalidFileException:
            print('Please, enter a valid file name')
            continue

        sheets = wb1.sheetnames
        print(sheets)
        choosing_sheet = input('Please, enter a sheet name, you wish to import')

        try:
            sheet = wb1[choosing_sheet]
        except KeyError:
            print('Please, enter a valid sheet name')
            continue

        rows = sheet.max_row
        cols = sheet.max_column

        name_of_fields = []
        fields_values = []
#  Print persons. This pattern doesn't matter for our task,
#  so it could be commented
        for i in range(1, rows + 1):
            value_row = []
            for j in range(1, cols + 1):
                cell = sheet.cell(row=i, column=j)
                value = str(cell.value)
                if i == 1:
                    name_of_fields.append(value)
                else:
                    value_row.append(value)
            if i != 1:
                fields_values.append(value_row)


# copy to file from other
        wb2 = openpyxl.load_workbook(work_file+'.xlsx')
        ws2 = wb2.active
        mr = sheet.max_row
        mc = sheet.max_column
        for i in range(1, mr + 1):
            for j in range(1, mc + 1):
                c = sheet.cell(row=i, column=j)
                ws2.cell(row=i, column=j).value = c.value
                dim_holder = DimensionHolder(worksheet=ws2)
                for col in range(ws2.min_column, ws2.max_column + 1):
                    dim_holder[get_column_letter(col)] = ColumnDimension(ws2, min=col, max=col, width=20)

                ws2.column_dimensions = dim_holder
                # wb2.save(work_file+'.xlsx')

        print(name_of_fields, end='\n')

        print(*fields_values, sep='\n')


    if ch == 0:
        break




class TestMethods(unittest.TestCase):

    def test_positive(self):
        ch = "ok"
        message = "Test value is  none."
        self.assertIsNotNone(ch, message)


    def test_int(self):
        ch = 2
        message = "given object is not instance of Myclass."
        self.assertIsInstance(ch, (str, int), message)


if __name__ == '__main__':
    unittest.main()